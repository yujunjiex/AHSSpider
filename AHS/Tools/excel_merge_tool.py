# coding: UTF-8

import openpyxl


def merge_same_rows(excel_path):
    """合并第一列相同内容的单元格(除表头外)"""
    wb = openpyxl.load_workbook(excel_path)

    for sheet_name in wb.sheetnames:
        print("正在合并{}".format(sheet_name))
        ws = wb.get_sheet_by_name(sheet_name)

        rows = list(ws.rows)
        row_count = len(rows)
        index = 1

        while index < row_count-1:  # 略过表头
            value = rows[index][0].value
            begin = index + 1
            end = index + 1

            for _row in rows[index + 1:]:
                # print(_row[0].value)
                if _row[0].value == value:
                    end += 1
                else:
                    break
            # 合并单元格
            ws.merge_cells(f'A{str(begin)}:A{str(end)}')
            index += 1

    wb.save(excel_path)


