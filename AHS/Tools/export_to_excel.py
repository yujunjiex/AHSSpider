# coding: UTF-8
"""将数据导出到excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pymongo
from Tools.excel_merge_tool import merge_same_rows

client = pymongo.MongoClient(host="127.0.0.1", port=27017)

# 字段
tablet_words = ('型号', '报价', "储存容量", "内存", "网络模式")
laptop__words = ('型号', '报价', "处理器", "机械硬盘", "固态硬盘", "显卡", "内存", "屏幕类型")
essential_words_dict = {
    "AHSTablet": tablet_words,
    "AHSLaptop": laptop__words
}

excel_name_dict = {
    "AHSTablet": "爱回收平板电脑报价",
    "AHSLaptop": "爱回收笔记本电脑报价"
}

tablet_excel_position = {
    "型号": 'A',
    "报价": 'B',
    "储存容量": 'C',
    "内存": 'D',
    "网络模式": 'E'
}

laptop_excel_position = {
    "型号": 'A',
    "报价": 'B',
    "处理器": 'C',
    "机械硬盘": 'D',
    "固态硬盘": 'E',
    "显卡": 'F',
    "内存": 'G',
    "屏幕类型": 'H'
}

position_dict = {
    "AHSTablet": tablet_excel_position,
    "AHSLaptop": laptop_excel_position
}


def write_mongodb_to_excel(database_name, excel_type="xlsx", path="./"):
    """把mongodb中的数据导出到excel"""
    db = client[database_name]
    # filename = path + excel_name_dict[database_name] + ".{}".format(excel_type)   # excel文件名
    filename = path   # excel文件名
    wb = Workbook()  # 创建一个workbook
    default_sheet = True
    aligment_center = Alignment(horizontal='center', vertical='center')
    for coll_name in db.list_collection_names(session=None):
        print("处理{}中".format(coll_name))
        if default_sheet is True:
            ws = wb.active   # 拿到默认表
            ws.title = coll_name
            default_sheet = False
        else:
            ws = wb.create_sheet(coll_name)  # 创建新表

        table_title = essential_words_dict[database_name]  # 创建表头
        for col in range(len(table_title)):
            c = col + 1
            cell = ws.cell(row=1, column=c)
            cell.value = table_title[col]
            cell.font = Font(name="微软雅黑", size=14)
            cell.alignment = aligment_center

        index = 2  # 从第二行开始添加数据

        table = db[coll_name]  # 拿到数据库表
        excel_position_dict = position_dict[database_name]  # 拿到字段位置对应字典
        # 设置行高列宽
        ws.row_dimensions[1].height = 24
        for value in excel_position_dict.values():
            ws.column_dimensions[value].width = 30

        for item in table.find(projection={"_id": 0, "id": 0}):
            for key in item:
                position = excel_position_dict[key]
                cell = ws[position+str(index)]
                cell.value = item[key]
                cell.font = Font(name="宋体", size=14)
                cell.alignment = aligment_center
            index += 1

    wb.save(filename)  # 保存文件


if __name__ == '__main__':
    write_mongodb_to_excel("AHSTablet", path=r'F:/爱回收数据备份/爱回收平板电脑报价9-28.xlsx')
    merge_same_rows("F:/爱回收数据备份/爱回收平板电脑报价9-28.xlsx")
    write_mongodb_to_excel("AHSLaptop", path=r'F:/爱回收数据备份/爱回收笔记本电脑报价9-28.xlsx')
    merge_same_rows("F:/爱回收数据备份/爱回收笔记本电脑报价9-28.xlsx")
